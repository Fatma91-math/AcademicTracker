import os
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


class Report:

    def __init__(self):

        self.report_folder = "reports"

        os.makedirs(
            self.report_folder,
            exist_ok=True
        )

    def create_report(self, source, old_profile, new_profile):

        workbook = Workbook()

        sheet = workbook.active

        sheet.title = source.upper()

        # -----------------------------
        # Başlıq
        # -----------------------------

        sheet["A1"] = "Academic Tracker Report"

        sheet["A1"].font = Font(
            bold=True,
            size=16
        )

        sheet["A2"] = "Source"

        sheet["B2"] = source.upper()

        sheet["A3"] = "Generated"

        sheet["B3"] = datetime.now().strftime(
            "%d %B %Y %H:%M:%S"
        )

        # -----------------------------

        sheet["A5"] = "Metric"
        sheet["B5"] = "Previous"
        sheet["C5"] = "Current"
        sheet["D5"] = "Change"

        header_fill = PatternFill(
            fill_type="solid",
            start_color="4F81BD",
            end_color="4F81BD"
        )

        for cell in sheet[5]:

            cell.font = Font(
                bold=True,
                color="FFFFFF"
            )

            cell.fill = header_fill

            cell.alignment = Alignment(
                horizontal="center"
            )

        # -----------------------------

        row = 6

        for key, value in new_profile.items():

            if key == "source":
                continue

            previous = ""

            if old_profile is not None:

                previous = old_profile.get(
                    key,
                    ""
                )

            current = value

            change = ""

            if isinstance(previous, (int, float)) and isinstance(current, (int, float)):

                diff = current - previous

                if diff > 0:
                    change = f"+{diff}"

                elif diff < 0:
                    change = str(diff)

                else:
                    change = "No Change"

            elif previous == current:

                change = "No Change"

            else:

                change = "Changed"

            sheet.cell(row=row, column=1).value = key
            sheet.cell(row=row, column=2).value = previous
            sheet.cell(row=row, column=3).value = current
            sheet.cell(row=row, column=4).value = change

            # -------------------------
            # Rəngləmə
            # -------------------------

            if isinstance(change, str):

                if change.startswith("+"):

                    sheet.cell(row=row, column=4).fill = PatternFill(
                        fill_type="solid",
                        start_color="C6EFCE",
                        end_color="C6EFCE"
                    )

                elif change.startswith("-"):

                    sheet.cell(row=row, column=4).fill = PatternFill(
                        fill_type="solid",
                        start_color="FFC7CE",
                        end_color="FFC7CE"
                    )

                elif change == "No Change":

                    sheet.cell(row=row, column=4).fill = PatternFill(
                        fill_type="solid",
                        start_color="D9D9D9",
                        end_color="D9D9D9"
                    )

            row += 1

        # -----------------------------
        # Sütun ölçüləri
        # -----------------------------

        sheet.column_dimensions["A"].width = 25
        sheet.column_dimensions["B"].width = 20
        sheet.column_dimensions["C"].width = 20
        sheet.column_dimensions["D"].width = 18

        filename = os.path.join(
            self.report_folder,
            f"{source}_report.xlsx"
        )

        workbook.save(filename)

        print(f"{filename} created successfully.")